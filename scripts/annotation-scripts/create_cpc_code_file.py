#!/usr/bin/python
from openpyxl import load_workbook
import re


# column 1 is SPECIES_CODE
# column 2 is CATAMI_DISPLAY_NAME
# column 3 is CATAMI_PARENT_ID

def remove_prefix(prefix, node):
    code_prefix = create_cpc_code(prefix)

    node.name = node.name.replace(prefix, "")
    node.cpc_code = node.cpc_code.replace(code_prefix, "", 1)

    for c in node.children:
        remove_prefix(prefix, c)

def split_catami_name(name):
    return [r.strip() for r in name.split(':')]

def create_cpc_code(name):
    # this does not guarantee uniqueness
    parts = re.split('\W+', name)
    code = ""
    for p in parts:
        if p != "":
            code += p[0].upper()

    return code

class Entry(object):
    def __init__(self, code, name, parent_code, cpc_code=None):
        self.code = code
        self.name = name
        self.parent_code = parent_code

        self.parent = None
        self.local_name = split_catami_name(name)[-1]
        self.default_code = create_cpc_code(self.local_name)
        self.cpc_code = cpc_code or create_cpc_code(name)
        self.children = []

    def __str__(self):
        child_string = "["
        for c in self.children:
            child_string += str(c)
        child_string += "]"
        return '\n<"{0} ({1})": parent: {2},children:\n{3}\n>'.format(self.name, self.code, self.parent_code, child_string)

    def __repr__(self):
        #child_string = "["
        #for c in self.children:
        #    child_string += str(c)
        #child_string += "]"
        return '\n<"{0} ({1})": parent: {2}, children:{3}>'.format(self.name, self.code, self.parent_code, len(self.children))

    def add_child(self, child, position=None):
        if not position is None:
            self.children.insert(position, child)
        else:
            self.children.append(child)

        child.parent = self

    def unique_cpc_code(self):
        code = ""

        # some hard coded values...
        # categories...
        if self.name == "Corals":
            return "C"
        elif self.name == "Tape, wand, shadow":
            return "TWS"

        # labels...
        if self.name == "Tape":
            return "TAPE"
        elif self.name == "Wand":
            return "WAND"
        elif self.name == "Shadow":
            return "SHAD"

        if self.parent:
            # get the parents unique code
            code += self.parent.unique_cpc_code()

            # get the siblings we need to compare against
            siblings = self.parent.children

            # check if the default name clashes with other siblings
            for s in siblings:
                if s is not self:
                    # check the default split
                    if s.default_code == self.default_code:
                        break
                    # or any codes of the same length...
                    if s.local_name.replace(" ","").upper()[0:len(self.default_code)] == self.default_code:
                        break
            else:
                # the default split is unique
                # so we can use that
                code += self.default_code
                return code

            # go through until we get a difference
            # if we are at this point we want to also make sure
            # we aren't comparing against cases where the default will
            # work...
            index = 1
            local_name = self.local_name.replace(" ", "").upper()
            for s in siblings:
                if s is not self:
                    other_name = s.local_name.replace(" ", "").upper()
                    while other_name[0:index] == local_name[0:index]:
                        index += 1


            code += self.local_name[0:index].upper()

        return code

if __name__ == '__main__':
    import argparse
    import logging
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook_file")
    parser.add_argument("output_file")
    parser.add_argument("--worksheet-name", default='Sheet1', help="Name of the sheet.")
    args = parser.parse_args()

    workbook = load_workbook(args.workbook_file)

    worksheet = workbook.get_sheet_by_name(args.worksheet_name)

    root = Entry("NO CODE", "Root", "NO CODE")

    loose_list = []
    leaf_list = []

    for data_row in xrange(1, worksheet.get_highest_row()):
        species_code = worksheet.cell(row=data_row, column=0).value
        display_name = worksheet.cell(row=data_row, column=1).value
        parent_code = worksheet.cell(row=data_row, column=2).value

        # discard first entry if it is a heading...
        if species_code == "SPECIES_CODE":
            continue

        new_entry = Entry(species_code, display_name, parent_code)
        if parent_code == None:
            root.add_child(new_entry)
            leaf_list.append(new_entry)
        else:
            loose_list.append(new_entry)


    # while we still have entries...
    incorrect_parent = []
    while len(leaf_list):
        entry = leaf_list.pop()
        # iterate backwords, permitting removal as we go...
        for i in xrange(len(loose_list) - 1, -1, -1):
            child_entry = loose_list[i]

            if child_entry.parent_code == entry.code:
                if not child_entry.name.startswith(entry.name) and not entry.name in ['Biota', 'Physical']:
                    logging.warning("Incorrect parent: %s has parent %s", child_entry.name, entry.name)
                    incorrect_parent.append(child_entry)
                entry.add_child(child_entry)
                del loose_list[i]
                leaf_list.append(child_entry)


    # we now have the full list worked out...
    # it reads the title row...
    if len(loose_list) > 1:
        #logging.warning("LOOSE LIST: %s", loose_list)
        for entry in loose_list:
            logging.error("No object matching parent code %s for %s", entry.parent_code, entry.name)

    #print "LOOSE LIST:"
    #for e in loose_list:
    #    print e

    # the two tree tops are:
    #
    # Biota 80000000
    # Physical 82000000

    # but they can be identified by their lack of parent
    # also Tubes, Wands, Shadows need to be in existence and moved around
    # and Corals need to be promoted to the top level categories

    #print "Top Level Nodes"
    #for e in root.children:
    #    print e.name

    try:
        biota = next(e for e in root.children if e.name == 'Biota')
        physical = next(e for e in root.children if e.name == 'Physical')
    except StopIteration:
        print "Could not find top level 'Biota' node."
        quit(-1)

    try:
        cnidaria = next(e for e in biota.children if e.name == 'Cnidaria')
        #print "Found Cnidaria"
        corals = next(e for e in cnidaria.children if e.name == 'Cnidaria: Corals')
        #print "Found Corals"
    except StopIteration:
        print "Could not find 'Cnidaria' and/or child 'Corals' nodes."
        quit(-1)

    try:
        substrate = next(e for e in physical.children if e.name == "Substrate")
    except StopIteration:
        print "Could not find 'substrate' under physical"
        quit(-1)

    # these need to exist for CPC
    tws = Entry("", "Tape, wand, shadow", "")
    tws.add_child(Entry("", "Tape", ""))
    tws.add_child(Entry("", "Wand", ""))
    tws.add_child(Entry("", "Shadow", ""))

    # move Corals up to be the first top level category
    biota.add_child(corals, position=0)
    # chuck substrate near the bottom
    biota.add_child(substrate)
    # place tws at the bottom of the categories
    biota.add_child(tws)
    biota.parent = None

    # remove coral from the children of cnidaria
    cnidaria.children.remove(corals)

    # remove cnidaria from corals
    # and all children
    remove_prefix("Cnidaria: ", corals)


    # add categories here that have special colours
    # otherwise they get the default
    colours = {
        "Coral": "FF0000",
    #    "Worms": "8F8F8F",
        "Sponges": "FFFF00",
    #    "Seagrasses":
    #    "Sea spiders":
    #    "Molluscs":
        "Macroalgae": "00FF40",
    #    "Jellies":
    #    "Fishes":
        "Echinoderms": "FF00FF",
    #    "Crustacea":
        "Cnidaria": "FF8080",
        "Bryozoa": "80FFFF",
    #    "Brachipods":
    #    "Bioturbation":
    #    "Bacterial mats":
        "Ascidian": "FF80FF",
        "Substrata": "DFDFDF",
    #    "Tape, wand, shadow":
    }


    # the header
    outfile = open(args.output_file, 'w')
    outfile.write("DFDFDF\nFFFFFF\n{0}\n".format(len(biota.children)))

    # now the categories
    for category in biota.children:
        name = category.name
        if name == "Corals":
            name = "Coral"
        colour = colours.get(name, "8F8F8F") # if not there use a default
        outfile.write('"{0}", "{1}", {2}\n'.format(category.unique_cpc_code(), name, colour))

    def print_child_codes(node, category):
        for c in node.children:
            if c.name in ["Tape", "Wand", "Shadow"]:
                outfile.write('"{0}", "{1}", "{2}"\n'.format(c.unique_cpc_code(), c.name, category))
            else:
                outfile.write('"{0}", "{1} ({3})", "{2}"\n'.format(c.unique_cpc_code(), c.name, category, c.code))
            print_child_codes(c, category)


    # now sub categories
    for c in biota.children:
        code = c.unique_cpc_code()
        outfile.write('"{0}", "{1} ({3})", "{2}"\n'.format(code + "1", c.name, code, c.code))
        print_child_codes(c, c.unique_cpc_code())


    def return_cpc_codes(node):
        codes = []
        for c in node.children:
            codes.extend(return_cpc_codes(c))
        codes.append(node.unique_cpc_code())
        return codes

    all_codes = return_cpc_codes(biota)

    seen = set()
    seen_add = seen.add

    duplicates = set(x for x in all_codes if x in seen or seen_add(x))

    if len(duplicates):
        for d in duplicates:
            objects = []
            for k,v in all_codes.iteritems():
                if d == v:
                    objects.append(k)
            logging.error("Code %s refers to multiple objects: %s", d, objects)

    for k in sorted(all_codes):
        if len(k) > 5:
            logging.warning("Long code %s.", k)



    # print tail end

    outfile.write(""""Blank", "Blank", "Blank"
NOTES,NOTES,NOTES
"D", "Dead", "NA"
"DISS", "Diseased sponge", "NA"
"MACS", "Macroalgae covering sponge", "NA"
"BL", "Bleached coral point", "NA"
"DIS", "Diseased coral point", "NA"
"LES", "Lesion", "NA"
"SED", "Sediment damage", "NA"
"STD", "Storm damage", "NA"
"ON1", "Other notes 1", "NA"
""")
    if len(duplicates):
        logging.error("Duplicate codes were generated")
        print "Duplicates:\n", list(duplicates)
